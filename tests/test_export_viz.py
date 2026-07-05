"""Tests for the timeline SVG renderer and the Excel statistics export (offline)."""
import sys
import zipfile
from io import BytesIO
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.analysis.export import statistics_to_xlsx  # noqa: E402
from app.correlation.timeline import render_timeline_svg  # noqa: E402

_TL = {
    "sessions": [
        {"ip": "45.155.205.108", "ua": "sqlmap/1.7", "start": "2026-07-01T10:15:01",
         "end": "2026-07-01T10:16:05", "event_count": 9, "n404": 4, "bytes_out": 0,
         "tags": ["sqli", "scanner_ua"], "phase": "exploitation"},
        {"ip": "45.155.205.108", "ua": "Mozilla/5.0", "start": "2026-07-01T11:05:00",
         "end": "2026-07-01T11:07:45", "event_count": 3, "n404": 0, "bytes_out": 41231882,
         "tags": ["webshell"], "phase": "post-exploitation"},
        {"ip": "203.0.113.10", "ua": "Chrome", "start": "2026-07-01T09:00:01",
         "end": "2026-07-01T09:05:20", "event_count": 4, "n404": 0, "bytes_out": 10000,
         "tags": [], "phase": "recon"},
    ],
    "attack_chains": [
        {"ip": "45.155.205.108", "first_seen": "2026-07-01T10:15:01",
         "phases": ["exploitation", "post-exploitation"], "sessions": 2},
        {"ip": "203.0.113.10", "first_seen": "2026-07-01T09:00:01",
         "phases": ["recon"], "sessions": 1},
    ],
}


def test_timeline_svg_basic():
    svg = render_timeline_svg(_TL)
    assert svg.startswith("<svg") and svg.rstrip().endswith("</svg>")
    assert svg.count("<rect") >= 3            # one bar per session (+legend swatches)
    assert "#ffa726" in svg and "#ef5350" in svg  # exploitation + post-exploitation colours
    assert "45.155.205.108" in svg and "203.0.113.10" in svg  # IP lane labels
    assert "<title>" in svg                   # hover tooltips


def test_timeline_svg_empty():
    assert render_timeline_svg({"sessions": []}) == ""
    assert render_timeline_svg(None) == ""


def test_timeline_svg_ignores_unparseable_times():
    tl = {"sessions": [{"ip": "1.2.3.4", "start": "not-a-date", "end": "x",
                        "phase": "recon", "tags": [], "event_count": 1}],
          "attack_chains": []}
    assert render_timeline_svg(tl) == ""


def _sample_stats():
    def rank(label, n):
        return [{"label": label, "count": n, "share": 100.0, "pct": 100.0}]
    return {
        "case_id": 1, "top_n": 20,
        "summary": {"total_events": 100, "unique_ips": 5, "unique_user_agents": 4,
                    "unique_urls": 20, "unique_domains": 2, "total_bytes": 1073741824,
                    "first_seen": "2026-07-01T09:00:00", "last_seen": "2026-07-01T12:00:00"},
        "top_ips": rank("1.2.3.4", 50), "top_ip_bytes": rank("1.2.3.4", 1048576),
        "top_user_agents": rank("curl", 10), "top_urls": rank("/a", 30),
        "top_domains": rank("x.com", 40), "top_referrers": rank("y.com", 5),
        "top_methods": rank("GET", 90), "top_status": rank(200, 80),
        "status_classes": [{"label": "2xx Success", "cls": "success", "count": 80, "share": 80.0},
                           {"label": "4xx Client error", "cls": "client_error", "count": 20, "share": 20.0}],
        "timeseries": [{"t": "2026-07-01 09:00", "count": 10},
                       {"t": "2026-07-01 10:00", "count": 40},
                       {"t": "2026-07-01 11:00", "count": 25}],
    }


def test_xlsx_export_is_valid_and_has_charts():
    data = statistics_to_xlsx(_sample_stats())
    assert isinstance(data, bytes) and data[:2] == b"PK"   # xlsx == zip

    zf = zipfile.ZipFile(BytesIO(data))
    names = zf.namelist()
    assert any(n.startswith("xl/charts/chart") for n in names), "no charts embedded in workbook"

    from openpyxl import load_workbook
    wb = load_workbook(BytesIO(data))
    for sheet in ("Summary", "Top Source IPs", "Top IPs by Bytes", "Status classes",
                  "Traffic over time"):
        assert sheet in wb.sheetnames, f"missing sheet {sheet}"
    assert wb["Top Source IPs"]["A2"].value == "1.2.3.4"


if __name__ == "__main__":
    fns = [v for k, v in sorted(globals().items()) if k.startswith("test_")]
    for fn in fns:
        fn()
        print(f"PASS {fn.__name__}")
    print(f"\nAll {len(fns)} export/viz tests passed.")
