"""Export case statistics to an Excel workbook (.xlsx) with native charts.

Native openpyxl charts (not images) mean the graphs render and stay editable in
Excel / LibreOffice, and the export works offline and inside the packaged exe.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

_HEADER = Font(bold=True)


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _ranking_sheet(wb, title, items, label_hdr="Item", value_hdr="Count",
                   as_mb=False):
    ws = wb.create_sheet(title[:31])
    ws.append([label_hdr, value_hdr, "Share %"])
    for c in ws[1]:
        c.font = _HEADER
    for it in items:
        val = round(it["count"] / 1048576, 2) if as_mb else it["count"]
        ws.append([str(it["label"]), val, it.get("share", 0)])
    _autosize(ws, [48, 16, 10])
    if items:
        chart = BarChart()
        chart.type = "bar"
        chart.title = title
        chart.legend = None
        chart.height = max(6, min(len(items) * 0.5 + 2, 18))
        chart.width = 20
        data = Reference(ws, min_col=2, min_row=1, max_row=len(items) + 1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(items) + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E2")
    return ws


def statistics_to_xlsx(stats: dict) -> bytes:
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws.append([f"Aura DFIR — Case #{stats['case_id']} statistics"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    s = stats["summary"]
    rows = [
        ("Total events", s["total_events"]),
        ("Unique source IPs", s["unique_ips"]),
        ("Unique user-agents", s["unique_user_agents"]),
        ("Unique URLs", s["unique_urls"]),
        ("Domains", s["unique_domains"]),
        ("Total bytes served", s["total_bytes"]),
        ("Data served (GB)", round(s["total_bytes"] / 1073741824, 3)),
        ("First event", s.get("first_seen") or ""),
        ("Last event", s.get("last_seen") or ""),
        ("Top-N shown", stats["top_n"]),
    ]
    ws.append(["Metric", "Value"])
    ws["A3"].font = _HEADER
    ws["B3"].font = _HEADER
    for r in rows:
        ws.append(list(r))
    _autosize(ws, [24, 40])

    # --- status class distribution (with chart) ---
    sc = wb.create_sheet("Status classes")
    sc.append(["Class", "Count", "Share %"])
    for c in sc[1]:
        c.font = _HEADER
    for row in stats["status_classes"]:
        sc.append([row["label"], row["count"], row["share"]])
    _autosize(sc, [22, 14, 10])
    if stats["status_classes"]:
        ch = BarChart()
        ch.title = "HTTP status classes"
        ch.legend = None
        ch.height, ch.width = 8, 16
        data = Reference(sc, min_col=2, min_row=1, max_row=len(stats["status_classes"]) + 1)
        cats = Reference(sc, min_col=1, min_row=2, max_row=len(stats["status_classes"]) + 1)
        ch.add_data(data, titles_from_data=True)
        ch.set_categories(cats)
        sc.add_chart(ch, "E2")

    # --- ranking sheets ---
    _ranking_sheet(wb, "Top Source IPs", stats["top_ips"], "Source IP")
    _ranking_sheet(wb, "Top IPs by Bytes", stats["top_ip_bytes"], "Source IP",
                   value_hdr="MB", as_mb=True)
    _ranking_sheet(wb, "Top User-Agents", stats["top_user_agents"], "User-agent")
    _ranking_sheet(wb, "Top URLs", stats["top_urls"], "URL")
    _ranking_sheet(wb, "Top Domains", stats["top_domains"], "Domain")
    _ranking_sheet(wb, "Top Referrers", stats["top_referrers"], "Referrer")
    _ranking_sheet(wb, "HTTP Methods", stats["top_methods"], "Method")
    _ranking_sheet(wb, "Top Status Codes", stats["top_status"], "Status")

    # --- traffic over time (line chart) ---
    series = stats.get("timeseries", [])
    tw = wb.create_sheet("Traffic over time")
    tw.append(["Time", "Requests"])
    for c in tw[1]:
        c.font = _HEADER
    for pt in series:
        tw.append([pt["t"], pt["count"]])
    _autosize(tw, [22, 14])
    if series:
        lc = LineChart()
        lc.title = "Requests over time"
        lc.legend = None
        lc.height, lc.width = 9, 24
        data = Reference(tw, min_col=2, min_row=1, max_row=len(series) + 1)
        cats = Reference(tw, min_col=1, min_row=2, max_row=len(series) + 1)
        lc.add_data(data, titles_from_data=True)
        lc.set_categories(cats)
        tw.add_chart(lc, "D2")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
